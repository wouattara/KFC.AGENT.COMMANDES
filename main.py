from __future__ import annotations
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import anthropic
import pandas as pd
import openpyxl
import pdfplumber
import io
import os
import re

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory="static"), name="static")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")


# ── Parseurs ───────────────────────────────────────────────────────────────────

def parse_excel(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[-1]]
    produits, lines = [], []
    for row in ws.iter_rows(max_row=300, values_only=True):
        if not row[0] or not isinstance(row[0], str): continue
        nom = row[0].strip()
        if nom in ["Produit","Date de passation de commande","Date de livraison",""]: continue
        try:
            qte = float(row[8]) if len(row) > 8 and row[8] is not None else 0.0
            produits.append({"nom": nom, "qte": qte})
            lines.append(f"{nom} | {qte}")
        except: pass
    return produits, "\n".join(lines)

def parse_csv_bon(content: bytes):
    text = content.decode("utf-8", errors="replace")
    df = pd.read_csv(io.StringIO(text), sep=None, engine="python", header=1, skiprows=[0])
    produits, lines = [], []
    for _, row in df.iterrows():
        try:
            nom = str(row.get("Designation article","")).strip()
            qte = float(row.get("Quantite", 0))
            if nom and nom != "nan":
                produits.append({"nom": nom, "qte": qte})
                lines.append(f"{nom} | {qte}")
        except: pass
    return produits, "\n".join(lines)

def parse_pdf_bon(content: bytes):
    """
    Parseur PDF pour bon de confirmation STEF (et formats similaires).
    Format attendu : Ligne | Article | Désignation | Temp. | Unité | Qté cdée | Qté Conf | Rupt
    """
    # Pattern ligne produit : numéro ligne, éventuellement X, code article 6 chiffres,
    # désignation, température, COL, qté commandée, qté confirmée, rupture
    pattern = re.compile(
        r'^(\d+)\s+(?:X\s+)?(\d{5,6})\s+(.+?)\s+(SEC|SURG|FRAIS)\s+COL\s+(\d+)\s+(\d+)\s+\d+\s*$'
    )

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    # Reconstituer les lignes fragmentées (désignation sur 2 lignes)
    raw_lines = full_text.split('\n')
    merged = []
    i = 0
    while i < len(raw_lines):
        line = raw_lines[i].strip()
        if line and i + 1 < len(raw_lines):
            nxt = raw_lines[i + 1].strip()
            # Si la ligne suivante est une suite de désignation (pas un n° de ligne ni en-tête)
            if nxt and not re.match(r'^\d{2,3}\s', nxt) and not re.match(
                r'^(Page|Ligne|Totaux|CONFIRMATION|Notre|Pour|STEF|KFC|Code|Commandé|Livré)', nxt
            ):
                line = line + ' ' + nxt
                i += 1
        merged.append(line)
        i += 1

    produits = []
    raw_lines_out = []
    seen_articles = set()

    for line in merged:
        m = pattern.match(line)
        if m:
            article = m.group(2)
            designation = m.group(3).strip()
            qte_conf = int(m.group(6))
            # Dédupliquer par code article (les lignes "X" sont des anciens tarifs)
            if article not in seen_articles:
                seen_articles.add(article)
                produits.append({"nom": designation, "qte": qte_conf, "article": article})
                raw_lines_out.append(f"{designation} | {qte_conf}")

    # Fallback texte brut si aucun produit parsé (PDF image ou format inconnu)
    if not produits:
        return [], full_text

    return produits, "\n".join(raw_lines_out)

def parse_excel_bon(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    produits, lines = [], []
    for row in ws.iter_rows(max_row=300, values_only=True):
        if not row[0]: continue
        try:
            nom = str(row[0]).strip()
            qte = float(row[1]) if len(row) > 1 and row[1] is not None else 0
            produits.append({"nom": nom, "qte": qte})
            lines.append(f"{nom} | {qte}")
        except: pass
    return produits, "\n".join(lines)


# ── Matching ───────────────────────────────────────────────────────────────────

# Mots vides : conditionnements, unités, prépositions
STOP = {
    'X','G','KG','ML','CL','L','PCS','PC','SAC','SACS','BTE','BOITE','BOITES',
    'COLIS','LOT','PACK','SACHET','SACHETS','PM','GM','TENDER','FORMAT',
    'DE','DU','LE','LA','LES','ET','EN','AU','AUX',
    'KRUNCHY','PREMAR','LIGHT','SANS','SUCRE','PET','CODE','UNIQUE',
}

def normalize(s: str) -> str:
    s = s.upper()
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def extract_dimensions(s: str) -> set:
    """
    Extrait les tokens discriminants de taille/contenance/quantité :
    33CL, 50CL, 33G, 500G, GRANDE, MOYENNE, PETITE, MINI, LARGE, etc.
    Ces tokens doivent correspondre entre commande et bon pour valider le match.
    """
    s = s.upper()
    dims = set()
    # Dimensions numériques : 33CL, 50CL, 700G, 500G, 2500G, 30CM...
    for m in re.finditer(r'\b(\d+(?:[.,]\d+)?)\s*(CL|ML|KG|G|CM|L|OZ)\b', s):
        dims.add(m.group(1).replace(',','.') + m.group(2))
    # Quantités conditionnement : x10pcs, x12pcs, x1000pcs, x24pcs
    for m in re.finditer(r'[Xx]\s*(\d+)\s*(?:PCS|pcs)?\b', s):
        dims.add('QTE' + m.group(1))
    # Mots de taille
    for mot in ['GRANDE','MOYEN','MOYENNE','PETITE','PETIT','LARGE','MAXI','BIG']:
        if re.search(r'\b' + mot + r'\b', s):
            dims.add(mot)
    return dims

def keywords(s: str) -> list:
    words = normalize(s).split()
    result = []
    for w in words:
        if w in STOP or w in ('-','–','/','&'): continue
        if re.match(r'^\d+$', w): continue
        # Tokens conditionnement : 6X18PCS, 3X24PCS, etc.
        if re.match(r'^\d+[X×]\d+.*$', w) or re.match(r'^\d+X$', w): continue
        # Dimensions seules déjà capturées : 30CM, 25CL, 50CL
        if re.match(r'^\d+[A-Z]+$', w): continue
        result.append(w)
    return result

def dimensions_compatibles(a: str, b: str) -> bool:
    """
    Vérifie que les dimensions discriminantes sont compatibles.
    Si les deux ont des dimensions et qu'elles diffèrent → incompatible.
    Si l'un n'a pas de dimension → compatible (dimension absente = pas discriminant).
    """
    dims_a = extract_dimensions(a)
    dims_b = extract_dimensions(b)
    if not dims_a or not dims_b:
        return True
    # Au moins une dimension commune (ou aucune contradiction)
    return bool(dims_a & dims_b) or not (dims_a - dims_b) or not (dims_b - dims_a)

def score_match(a: str, b: str) :
    """
    Retourne (score, dimensions_ok).
    Le score seul ne suffit pas : on vérifie aussi la compatibilité des dimensions.
    """
    kw_a = keywords(a)
    kw_b = keywords(b)
    if not kw_a or not kw_b:
        return 0.0, False

    set_a, set_b = set(kw_a), set(kw_b)
    inter = set_a & set_b
    union = set_a | set_b

    # Jaccard
    jaccard = len(inter) / len(union) if union else 0

    # Containment : tous les mots du plus court sont dans le plus long
    shorter = set_a if len(set_a) <= len(set_b) else set_b
    containment = len(inter) / len(shorter) if shorter else 0

    # Préfixe : 2 premiers mots-clés du plus court dans le plus long
    kw_short = kw_a if len(kw_a) <= len(kw_b) else kw_b
    kw_long  = kw_b if len(kw_a) <= len(kw_b) else kw_a
    pfx = min(2, len(kw_short))
    prefix_score = sum(1 for w in kw_short[:pfx] if w in set(kw_long)) / pfx if pfx else 0

    score = max(jaccard, containment * 0.88, prefix_score * 0.92)
    dims_ok = dimensions_compatibles(a, b)

    return score, dims_ok

SEUIL_OK       = 0.72
SEUIL_APPROCHE = 0.55

def letter_products(commande: list, bon: list) -> dict:
    bon_index = {normalize(p["nom"]): p for p in bon}
    matched_bon = set()
    lettrage = []

    for item in commande:
        norm_cmd = normalize(item["nom"])
        qte_cmd  = item["qte"]

        # Chercher le meilleur match avec contrainte de dimensions
        best_score, best_key, best_item = 0.0, None, None
        for bk, bp in bon_index.items():
            s, dims_ok = score_match(norm_cmd, bk)
            # Pénaliser fortement si dimensions incompatibles
            if not dims_ok:
                s *= 0.4
            if s > best_score:
                best_score, best_key, best_item = s, bk, bp

        if best_score >= SEUIL_OK:
            matched_bon.add(best_key)
            qte_rec = best_item["qte"]
            ecart   = qte_rec - qte_cmd
            statut  = "OK" if abs(ecart) < 0.01 else "ECART_QTE"
            lettrage.append({
                "produit_commande": item["nom"],
                "produit_bon":      best_item["nom"],
                "qte_commandee":    int(qte_cmd),
                "qte_recue":        int(qte_rec),
                "ecart":            int(ecart),
                "statut":           statut,
                "score":            round(best_score, 2),
            })
        elif best_score >= SEUIL_APPROCHE:
            matched_bon.add(best_key)
            qte_rec = best_item["qte"]
            ecart   = qte_rec - qte_cmd
            lettrage.append({
                "produit_commande": item["nom"],
                "produit_bon":      best_item["nom"],
                "qte_commandee":    int(qte_cmd),
                "qte_recue":        int(qte_rec),
                "ecart":            int(ecart),
                "statut":           "APPROCHE",
                "score":            round(best_score, 2),
            })
        else:
            statut = "NON_PASSE" if qte_cmd == 0 else "MANQUANT"
            lettrage.append({
                "produit_commande": item["nom"],
                "produit_bon":      None,
                "qte_commandee":    int(qte_cmd) if qte_cmd else 0,
                "qte_recue":        None,
                "ecart":            None,
                "statut":           statut,
                "score":            0,
            })

    # Produits du bon non matchés
    for bk, bp in bon_index.items():
        if bk not in matched_bon:
            lettrage.append({
                "produit_commande": None,
                "produit_bon":      bp["nom"],
                "qte_commandee":    None,
                "qte_recue":        int(bp["qte"]),
                "ecart":            None,
                "statut":           "NON_COMMANDE",
                "score":            0,
            })

    stats = {
        "total_commande": sum(1 for l in lettrage if l["produit_commande"]),
        "total_bon":      len(bon),
        "ok":             sum(1 for l in lettrage if l["statut"] == "OK"),
        "ecart_qte":      sum(1 for l in lettrage if l["statut"] == "ECART_QTE"),
        "approche":       sum(1 for l in lettrage if l["statut"] == "APPROCHE"),
        "manquant":       sum(1 for l in lettrage if l["statut"] == "MANQUANT"),
        "non_commande":   sum(1 for l in lettrage if l["statut"] == "NON_COMMANDE"),
        "non_passe":      sum(1 for l in lettrage if l["statut"] == "NON_PASSE"),
    }
    return {"lettrage": lettrage, "stats": stats}


# ── Synthèse IA ────────────────────────────────────────────────────────────────

def build_synthese_ia(result: dict) -> str:
    if not ANTHROPIC_API_KEY:
        return "⚠️ Clé API Anthropic non configurée."
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    s = result["stats"]
    issues = [l for l in result["lettrage"] if l["statut"] not in ("OK","NON_PASSE")]
    issues_txt = "\n".join(
        f"- [{l['statut']}] cmd='{l['produit_commande'] or ''}' / bon='{l['produit_bon'] or ''}' "
        f"(commandé:{l['qte_commandee']}, reçu:{l['qte_recue']}, écart:{l['ecart']})"
        for l in issues
    )
    prompt = f"""Tu es expert en gestion opérationnelle KFC franchisé.

Lettrage commande vs bon de validation :
- {s['total_commande']} produits commandés / {s['total_bon']} sur le bon
- {s['ok']} lettrés OK ✓ | {s['approche']} à vérifier | {s['ecart_qte']} écarts qté | {s['manquant']} manquants | {s['non_commande']} non commandés

INCOHÉRENCES :
{issues_txt or 'Aucune.'}

Synthèse opérationnelle concise pour le manager. 3 parties :
### Bilan global
### Points d'attention prioritaires  
### Actions recommandées
Maximum 8 lignes."""
    msg = client.messages.create(model="claude-sonnet-4-6", max_tokens=600,
                                  messages=[{"role":"user","content":prompt}])
    return msg.content[0].text


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index():
    with open("templates/index.html","r",encoding="utf-8") as f: return f.read()

@app.post("/analyze")
async def analyze(commande: UploadFile = File(...), bon: UploadFile = File(...)):
    try:
        cb = await commande.read()
        bb = await bon.read()
        ext_c = commande.filename.lower().split(".")[-1]
        if ext_c in ("xlsx","xls"): cp, cr = parse_excel(cb)
        elif ext_c == "csv":
            df = pd.read_csv(io.BytesIO(cb), sep=None, engine="python", header=0)
            cp = [{"nom":str(r.iloc[0]),"qte":float(r.iloc[1])} for _,r in df.iterrows() if len(r)>=2]
            cr = df.to_string()
        else: raise HTTPException(400,"Format commande non supporté")
        ext_b = bon.filename.lower().split(".")[-1]
        if ext_b == "pdf": bp, br = parse_pdf_bon(bb)
        elif ext_b == "csv": bp, br = parse_csv_bon(bb)
        elif ext_b in ("xlsx","xls"): bp, br = parse_excel_bon(bb)
        else: raise HTTPException(400,"Format bon non supporté")
        result = letter_products(cp, bp if bp else [])
        result["synthese"] = build_synthese_ia(result)
        return JSONResponse(result)
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, f"Erreur: {str(e)}")

@app.get("/health")
async def health():
    return {"status":"ok","api_key_set":bool(ANTHROPIC_API_KEY)}

@app.get("/version")
async def version():
    return {"version": "2.1-pdf-stef", "pdf_parser": "stef_pattern_v2"}
